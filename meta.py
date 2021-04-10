# Charles Clark 4/10/21 Program to get all keys in a JSON file. 
# Adds a new sheet in Excel document labeled by JSON file name.

import json
import argparse
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook

#get file from command line argument
parser = argparse.ArgumentParser()                                               
parser.add_argument("--file", "-f", type=str, required=True)
args = parser.parse_args()

#check if Excel file exists 
if os.path.isfile("udlmetadata.xlsx") and os.access("udlmetadata.xlsx", os.R_OK):
  wb = load_workbook("udlmetadata.xlsx")
else:
  wb = Workbook()

# read file
with open(args.file, 'r') as myfile:
  data=myfile.read()

# parse file int object
obj = json.loads(data)

# Check to see if file is already ran and has a sheet
# Rewrites if exists, creates if not
try:
 ws1 = wb[args.file]
except KeyError:
 ws1 = wb.create_sheet(args.file)

# Put each key in a new A cell
cell = 1
for x in obj[1]:
 print(x)
 ws1['A' + str(cell)] = str(x)
 cell += 1

# Save the file
wb.save("udlmetadata.xlsx")




